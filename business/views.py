from django.shortcuts import render
from django.http import HttpResponseRedirect,Http404
from django.urls import reverse
from .forms import BankRegistrationForm,UploadFileForm
from django import forms
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate
from django.contrib.auth import login as auth_login
from rolepermissions.roles import assign_role
from CreditSols.roles import Banker
from rolepermissions.checkers import has_role
from django.contrib.auth import get_user_model
from .models import BusinessAccount
from django.db import transaction
from personal.models import PersonalAccount,PersonalInfo,DuesInfo
import openpyxl
from datetime import datetime
from os import listdir
from engine.models import Score



# Create your views here.
current_month = (datetime.now()).strftime("%B")

@login_required
def dashboard(request):
    user = request.user
    if has_role(user,Banker):
        user_list = Score.objects.order_by("user")
        total_users = list(range(len(user_list)+1))
        total_users.remove(0)
        user_score = []
        user_contact = []
        user_emails = []
        user_names = []
        for user in user_list:
            userr = (get_user_model().objects.all().filter(username = user))
            email = userr.values("email")[0]["email"]
            fname = userr.values("first_name")[0]["first_name"]
            lname = userr.values("last_name")[0]["last_name"]
            username = (PersonalAccount.objects.all().filter(user = userr[0]))[0]
            contact = (PersonalInfo.objects.all().filter(user = username)).values("phone")[0]["phone"]
            name = fname + " " + lname
            score = user.credit_score
            user_score.append(score)
            user_names.append(name)
            user_emails.append(email)
            user_contact.append(contact)
        return render(request,'bank_dashboard.html',{'user_score':user_score,
                                                     'user_contact':user_contact,
                                                     'user_emails':user_emails,
                                                     'user_names':user_names,
                                                     'total_users':total_users})
    else:
        raise Http404


@login_required
def upload(request):
    user = request.user
    if has_role(user,Banker):
        if request.method == 'POST':
            form = UploadFileForm(request.POST, request.FILES)
            if form.is_valid():
                banker = BusinessAccount.objects.all().filter(user=request.user)[0]
                file = form.save(commit=False)
                file.user = banker
                file.save()
                dir_path = 'media/CIR/{0}/{1}'.format(banker, current_month)
                for f in listdir(dir_path):
                     if f.endswith('.' + 'xlsx'):
                         file_name = f
                file_path = '{0}/{1}'.format(dir_path,file_name)
                update_duesinfo(file_path,banker)
                messages.success(request,('File is sucessfully uploaded'))
                return HttpResponseRedirect(reverse('business:upload'))
            else:
                print(form.errors)
                messages.error(request,(form.errors))
        else:
            form = UploadFileForm()
        return render(request,'upload.html',{'form': form})
    else:
        raise Http404


def signup(request):
    if request.method == 'POST':
        bank_form = BankRegistrationForm(data=request.POST)

        if bank_form.is_valid()and bank_form.cleaned_data['password1'] == bank_form.cleaned_data['password2']:
            user = bank_form.save()
            user.set_password(bank_form.cleaned_data['password1'])
            user.save()
            t = BusinessAccount(user = user)
            t.save()
            assign_role(user,Banker)
            messages.success(request,('Registration complete'))
            return HttpResponseRedirect(reverse('business:signup'))

        elif bank_form.data['password1'] != bank_form.data['password2']:
            messages.error(request,('Passwords do not match'))

        else:
            messages.error(request,('bank_form.errors'))
    else:
        if request.user.is_authenticated:
            return HttpResponseRedirect(reverse('business:dashboard'))
        else:
            bank_form = BankRegistrationForm()
    return render(request,'accounts/bank_registration.html',
                         {'bank_form':bank_form,})


def login(request):
    if request.method == 'POST':
        name = (request.POST.get('name')).upper()
        username = (get_user_model().objects.all().filter(first_name = name)).values("username")[0]["username"]
        password = request.POST.get('password')
        user = authenticate(username=username, password=password)
        if user:
            if user.is_active and has_role(user,Banker):
                auth_login(request,user)
                return HttpResponseRedirect(reverse('business:dashboard'))
            else:
                messages.error(request,"Your account is not active")
                return render(request,'accounts/bank_login.html')
        else:
            messages.error(request,"Invalid Username or Password")
            return render(request,'accounts/bank_login.html')
    else:
        if request.user.is_authenticated:
            return HttpResponseRedirect(reverse('business:dashboard'))
        else:
            return render(request,'accounts/bank_login.html')


@transaction.atomic
def update_duesinfo(file_path,banker):
    all_pan_numbers = []
    reg_pan_numbers = []
    reg_users = []
    reg_data = []
    wb_obj = openpyxl.load_workbook(file_path)
    sheet_obj = wb_obj.active
    max_rows = sheet_obj.max_row
    rows_iter = sheet_obj.iter_rows(min_col = 4, min_row = 2, max_col = 13, max_row = max_rows)
    all_values = [[cell.value for cell in row] for row in rows_iter]
    for i in range(2, max_rows + 1):
        cell_obj = sheet_obj.cell(row = i, column = 4)
        all_pan_numbers.append(cell_obj.value)
    pan_numbers = PersonalInfo.objects.values_list('pan_number', flat=True)
    for pan in all_pan_numbers:
        if pan in pan_numbers:
            reg_pan_numbers.append(pan)
        else:
            continue
    for reg_pan in reg_pan_numbers:
        user_id = PersonalInfo.objects.all().filter(pan_number = reg_pan).values("user_id")[0]['user_id']
        username = (get_user_model().objects.all().filter(id = user_id)).values("username")[0]["username"]
        reg_users.append(username)
    for reg_pan in reg_pan_numbers:
        for value in all_values:
            if (reg_pan == value[0]):
                value.remove(reg_pan)
                reg_data.append(value)
    user_dict = dict(zip(reg_users,reg_data))
    for users in user_dict.keys():
        val = user_dict.get(users)
        userr = (get_user_model().objects.all().filter(username = users))[0]
        username = (PersonalAccount.objects.all().filter(user = userr))[0]
        DuesInfo.objects.create(user=username,bank=banker)
        identity = DuesInfo.objects.filter(user = username, bank = banker).values("id")[0]['id']
        data = DuesInfo.objects.get(id=identity)
        data.Currently_Owned_Amount = val[0]
        data.Previously_Owned_Amount = val[1]
        data.Total_Number_Of_Overdues_in_Current_Loans = val[2]
        data.Total_Overdue_Amount_of_Current_Loans = val[3]
        data.Maximum_Number_of_Days_Amount_is_Overdue_in_Current_Loans = val[4]
        data.Status = val[5][0]
        data.Total_Overdue_Amount_of_Current_Loans = val[6]
        data.Total_Overdue_Amount_of_Completed_Loans = val[7]
        data.Maximum_Number_of_Days_Amount_is_Overdue_in_Completed_Loans = val[8]
        data.save()
